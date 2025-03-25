import random
import time
from openpyxl import load_workbook


def main():
	#set up excel
	wb = load_workbook('cmsc441_paper1_data.xlsx')

	insertion = wb['insertion_sort_data']
	selection = wb['selection_sort_data']
	merge = wb['merge_sort_data']
	quick = wb['quick_sort_data']

	sizes = [100, 500, 1000, 3000, 5000, 7000, 10000, 14000, 20000, 50000] #n
	# sizes = [100000]
	count = 2 #row counter in excel

	for size in sizes: #loops through each preset n size
		for i in range(10): #10 trials for each n size
			arr = []
			for x in range(size): #populate list
				arr.append(random.randint(0, size))
			
			start = time.time() #insertion sort
			insertionSort(arr.copy(), size)
			end = time.time()
			insertion.cell(row = count, column = 6, value = size)
			insertion.cell(row = count, column = 7, value = (end - start) * 1000)
			wb.save('cmsc441_paper1_data.xlsx')

			start = time.time() #selection sort
			selectionSort(arr.copy(), size)
			end = time.time()
			selection.cell(row = count, column = 6, value = size)
			selection.cell(row = count, column = 7, value = (end - start) * 1000)
			wb.save('cmsc441_paper1_data.xlsx')

			start = time.time() #merge sort
			mergeSort(arr.copy(), 0, size - 1)
			end = time.time()
			merge.cell(row = count, column = 6, value = size)
			merge.cell(row = count, column = 7, value = (end - start) * 1000)
			wb.save('cmsc441_paper1_data.xlsx')

			start = time.time() #quick sort
			quickSort(arr.copy(), 0, size - 1)
			end = time.time()
			quick.cell(row = count, column = 6, value = size)
			quick.cell(row = count, column = 7, value = (end - start) * 1000)
			wb.save('cmsc441_paper1_data.xlsx')

			count += 1
			
	wb.save('cmsc441_paper1_data.xlsx')
		
def insertionSort(arr, size):
	for i in range(size):
		key = arr[i]
		j = i - 1
		#move elements greater than the key by 1
		while j >= 0 and key < arr[j]:
			arr[j + 1] = arr[j]
			j -= 1
		arr[j + 1] = key

def selectionSort(arr, size):
	for i in range(size):
		minIndex = i
		#find min element
		for j in range(i + 1, size):
			if arr[j] < arr[minIndex]:
				minIndex = j
		arr[i], arr[minIndex] = arr[minIndex], arr[i] #swap min element with the ith element


def mergeSort(arr, l, r):
	#recursevely split the array into halves
	if l < r:
		m = (l + r) // 2
		mergeSort(arr, l, m) #sort left half
		mergeSort(arr, m + 1, r) #sort right half
		merge(arr, l, m, r)

def merge(arr, l, m, r):
	#sizes of the two halves
	h1 = m - l + 1
	h2 = r - m

	L = [0] * h1
	R = [0] * h2

	#copy arrays to temp arrays
	for i in range(h1):
		L[i] = arr[l + i]
	for j in range(h2):
		R[j] = arr[m + 1 + j]

	i = 0
	j = 0
	k = l

	#merger arrays together
	while i < h1 and j < h2:
		if L[i] <= R[j]:
			arr[k] = L[i]
			i += 1
		else:
			arr[k] = R[j]
			j += 1
		k += 1

	#copy remaining elements
	while i < h1:
		arr[k] = L[i]
		i += 1
		k += 1
	
	while j < h2:
		arr[k] = R[j]
		j += 1
		k += 1

def quickSort(arr, low, high):
	if low < high:
		pivot = partition(arr, low, high) #index to partition
		
		quickSort(arr, low, pivot - 1) #before pivot
		quickSort(arr, pivot + 1, high) #after pivot
	return arr

def partition(arr, low, high):
	pivot = arr[high]
	i = low - 1

	#traverse through the array and move smaller elments to the left
	for j in range(low, high):
		if arr[j] < pivot:
			i += 1
			swap(arr, i, j)

	swap(arr, i + 1, high)
	return i + 1

def swap(arr, i, j): #swap two elements
	arr[i], arr[j] = arr[j], arr[i]

if __name__ == "__main__":
	main()